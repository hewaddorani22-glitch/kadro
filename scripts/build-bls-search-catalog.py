#!/usr/bin/env python3
"""Build the compact bilingual search snapshot from the official BLS 4.0 file.

Usage:
  python scripts/build-bls-search-catalog.py /path/to/BLS_4_0_Daten_2025_DE.xlsx

The workbook is deliberately not committed. It is available under CC BY 4.0
from https://blsdb.de/download. The generated module contains only the fields
the app needs for search: source code, two names and five nutrients per 100 g.
"""

from __future__ import annotations

import hashlib
import json
import sys
from pathlib import Path

from openpyxl import load_workbook


EXPECTED_SOURCE_SHA256 = "524bbefe25b691f5cb3de7a9f3e27fa2967aebfeabf217d99414ba7806e78c60"
EXPECTED_ROWS = 7140
HEADERS = {
    "code": "BLS Code",
    "de": "Lebensmittelbezeichnung",
    "en": "Food name",
    "calories": "ENERCC Energie (Kilokalorien) [kcal/100g]",
    "protein": "PROT625 Protein (Nx6,25) [g/100g]",
    "fat": "FAT Fett [g/100g]",
    "carbs": "CHO Kohlenhydrate, verfügbar [g/100g]",
    "fiber": "FIBT Ballaststoffe, gesamt [g/100g]",
}


def number(value: object) -> float | int:
    """BLS uses <LOD for a value below the detection limit; zero is its display value."""
    if value is None:
        return 0
    if isinstance(value, str):
        cleaned = value.strip()
        if not cleaned or cleaned == "-" or cleaned.startswith("<"):
            return 0
        value = cleaned
    try:
        parsed = round(float(value), 3)
    except (TypeError, ValueError):
        # BLS marker values such as TR (trace) are below a useful display
        # precision for Kandro's whole-number meal estimates.
        return 0
    return int(parsed) if parsed.is_integer() else parsed


def main() -> None:
    if len(sys.argv) != 2:
        raise SystemExit("Pass the BLS_4_0_Daten_2025_DE.xlsx path.")

    source = Path(sys.argv[1]).expanduser().resolve()
    digest = hashlib.sha256(source.read_bytes()).hexdigest()
    if digest != EXPECTED_SOURCE_SHA256:
        raise SystemExit(
            "The BLS workbook differs from the reviewed 2025 snapshot. "
            f"Expected {EXPECTED_SOURCE_SHA256}, got {digest}."
        )

    workbook = load_workbook(source, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    header = list(next(rows))
    columns = {name: header.index(label) for name, label in HEADERS.items()}

    output_rows = []
    for row in rows:
        output_rows.append([
            str(row[columns["code"]]).strip(),
            str(row[columns["de"]]).strip(),
            str(row[columns["en"]]).strip(),
            number(row[columns["calories"]]),
            number(row[columns["protein"]]),
            number(row[columns["carbs"]]),
            number(row[columns["fat"]]),
            number(row[columns["fiber"]]),
        ])

    if len(output_rows) != EXPECTED_ROWS:
        raise SystemExit(f"Expected {EXPECTED_ROWS} BLS rows, got {len(output_rows)}.")
    if len({row[0] for row in output_rows}) != len(output_rows):
        raise SystemExit("BLS codes are not unique.")

    target = Path(__file__).resolve().parents[1] / "supabase/functions/_shared/bls-search-data.mjs"
    preamble = f'''/**
 * GENERATED FILE — do not edit by hand.
 *
 * Source: Max Rubner-Institut, Bundeslebensmittelschlüssel 4.0 (2025), CC BY 4.0
 * DOI: 10.25826/Data20251217-134202-0
 * Workbook SHA-256: {digest}
 * Generator: scripts/build-bls-search-catalog.py
 *
 * Row shape: [code, German name, English name, kcal, protein, carbs, fat, fiber]
 * All nutrient values are per 100 g edible portion.
 */
export const BLS_SEARCH_WORKBOOK_SHA256 = '{digest}';
export const BLS_SEARCH_ROWS = Object.freeze([
'''
    body = ",\n".join(json.dumps(row, ensure_ascii=False, separators=(",", ":")) for row in output_rows)
    target.write_text(f"{preamble}{body}\n]);\n", encoding="utf-8")
    print(f"Generated {len(output_rows)} bilingual BLS foods at {target}.")


if __name__ == "__main__":
    main()
